import gc
import csv
import logging
from io import StringIO, BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import datetime, timedelta, date, time
from openpyxl import load_workbook
from fastapi import APIRouter, HTTPException, UploadFile, File, Depends
from app.crud import egat_api
from app.crud import electric as crud
from app.db.session import get_db
from app.schemas.utils import Items, Item, Msg, ItemWithPercent, TimeseriesItem, ItemWithTimestamp, LocationItem
from app.schemas import electric as schemas
from typing import Optional
from time import time as runtime
from math import floor

router = APIRouter()

logger = logging.getLogger(__name__)


@router.get("/current/supply")
async def get_current_supply(is_include_ips: Optional[bool] = True, source: int = 1, db: AsyncSession = Depends(get_db)) -> Items:
    start_time = runtime()
    items = Items(datetime=datetime.now(), status='ok')
    total_item = Item(tag='total')
    egat_item = Item(tag='egat')
    ipp_item = Item(tag='ipp')
    spp_item = Item(tag='spp')
    vspp_item = Item(tag='vspp')
    ips_item = Item(tag='ips')
    imp_item = Item(tag='imp')
    try:
        #* EGAT, IPP, SPP
        if source == 1:
            for data in egat_api.get_current_genMw_source1():
                if 'EGAT' in data['plant_type']:
                    egat_item.value += data['value']
                elif 'IPP' in data['plant_type']:
                    ipp_item.value += data['value']
                elif 'SPP' in data['plant_type']:
                    spp_item.value += data['value']
                elif 'imp' in data['plant_type']:
                    imp_item.value += data['value']
                items.datetime = data['data_timestamp']
        elif source == 2:
            for data in egat_api.get_current_genMw_source2():
                if 'EGAT' in data['plant_type']:
                    egat_item.value += data['value']
                elif 'IPP' in data['plant_type']:
                    ipp_item.value += data['value']
                elif 'SPP' in data['plant_type']:
                    spp_item.value += data['value']
                elif 'imp' in data['plant_type']:
                    imp_item.value += data['value']
                items.datetime = data['data_timestamp']   
        elif source == 3:
            for data in egat_api.get_current_genMw_source3():
                if data['COMPANYTYPE'] == 'EGAT':
                    egat_item.value += data['VALUE']
                elif data['COMPANYTYPE'] == 'IPP':
                    ipp_item.value += data['VALUE']
                elif 'SPP' in data['COMPANYTYPE']:
                    spp_item.value += data['VALUE']
            items.datetime = datetime.fromisoformat(data['TIMESTAMP'])
        else:
            raise HTTPException(status_code=400, detail='soucre should be 1, 2, or 3')
        
        min_today = items.datetime.hour*60
        # *VSPP
        data = egat_api.get_reqMw_selected_min(13, datetime.today(), min_today)
        vspp_item.value += data[1]
        data = egat_api.get_reqMw_selected_min(14, datetime.today(), min_today)
        vspp_item.value += data[1]
        data = egat_api.get_reqMw_selected_min(15, datetime.today(), min_today)
        vspp_item.value += data[1]
        data = egat_api.get_reqMw_selected_min(16, datetime.today(), min_today)
        vspp_item.value += data[1]
        data = egat_api.get_reqMw_selected_min(12, datetime.today(), min_today)
        vspp_item.value += data[1]

        # *IPS
        if is_include_ips:
            async for data in crud.get_latest_dummy_data(db=db, category='ips', data_timestamp=items.datetime):
                ips_item.value += data['value']

        total_item.value = egat_item.value + ipp_item.value + spp_item.value + vspp_item.value + ips_item.value + imp_item.value
        total_item.value = round(total_item.value, 4)
        egat_item.value = round(egat_item.value, 4)
        ipp_item.value = round(ipp_item.value, 4)
        spp_item.value = round(spp_item.value, 4)
        vspp_item.value = round(vspp_item.value, 4)
        ips_item.value = round(ips_item.value, 4)
        imp_item.value = round(imp_item.value, 4)
        items.items = [total_item, egat_item, ipp_item, spp_item, vspp_item, ips_item, imp_item]
    except Exception as e:
        logger.exception(e)
        items.status = 'error'
    logger.info(f'runtime: {runtime() - start_time:3f} s')
    return items

@router.get("/current/demand")
async def get_current_demand(
    is_include_ips: Optional[bool] = False,
    db: AsyncSession = Depends(get_db)
) -> Items:
    start_time = runtime()
    try:
        # *VSPP
        data = egat_api.get_reqMw(13, datetime.today())
        item_vspp_rcc1 = Item(tag='vspp_rcc1',
                                      value=data['list'][-1][-1])
        data = egat_api.get_reqMw(14, datetime.today())
        item_vspp_rcc2 = Item(tag='vspp_rcc2',
                                      value=data['list'][-1][-1])
        data = egat_api.get_reqMw(15, datetime.today())
        item_vspp_rcc3 = Item(tag='vspp_rcc3',
                                      value=data['list'][-1][-1])
        data = egat_api.get_reqMw(16, datetime.today())
        item_vspp_rcc4 = Item(tag='vspp_rcc4',
                                      value=data['list'][-1][-1])
        data = egat_api.get_reqMw(12, datetime.today())
        item_vspp_mcc = Item(tag='vspp_mcc',
                                     value=data['list'][-1][-1])

        # * Export
        data = egat_api.get_reqMw(6, datetime.today())
        item_exp_edl = Item(tag='exp_edl', value=data['list'][-1][-1])
        data = egat_api.get_reqMw(7, datetime.today())
        item_exp_tnp = Item(tag='exp_tnp', value=data['list'][-1][-1])

        # *3E
        data = egat_api.get_reqMw(1, datetime.today())
        item_rcc1 = Item(tag='rcc1', value=data['list'][-1][-1])
        data = egat_api.get_reqMw(2, datetime.today())
        item_rcc2 = Item(tag='rcc2', value=data['list'][-1][-1])
        data = egat_api.get_reqMw(3, datetime.today())
        item_rcc3 = Item(tag='rcc3', value=data['list'][-1][-1])
        data = egat_api.get_reqMw(4, datetime.today())
        item_rcc4 = Item(tag='rcc4', value=data['list'][-1][-1])
        data = egat_api.get_reqMw(5, datetime.today())
        item_mcc = Item(tag='mcc', value=data['list'][-1][-1])

        base_date = datetime.strptime(data['day'], "%d-%m-%Y")
        time_delta = timedelta(seconds=data['list'][-1][0])
        data_datetime = base_date + time_delta
        del data
        gc.collect()

        mea_value = item_mcc.value + item_vspp_mcc.value
        pea_value = item_rcc1.value +  item_rcc2.value + item_rcc3.value + item_rcc4.value
        pea_vspp_value = item_vspp_rcc1.value + item_vspp_rcc2.value + item_vspp_rcc3.value + item_vspp_rcc4.value
        pea_value += pea_vspp_value
        exp_value = item_exp_edl.value + item_exp_tnp.value
        total_value = mea_value + pea_value + exp_value

        # *ips dummy
        ips_value = 0
        if is_include_ips:
            async for ips_data in crud.get_latest_dummy_data_grouped_by_zone(
                    db=db, category='ips', data_timestamp=data_datetime):
                submit_ts = ips_data['submit_timestamp'].strftime("%Y-%m-%d %H:%M")
                data_ts = ips_data['data_timestamp'].strftime("%Y-%m-%d %H:%M")
                logger.debug(f'IPS {submit_ts}: {data_ts} {ips_data['zone']} {ips_data['value']}')
                if ips_data['zone'] == '':
                    mea_value += ips_data['value']
                else:
                    pea_value += ips_data['value']
                total_value += ips_data['value']
                ips_value += ips_data['value']


        # *EGAT customer direct
        egat_value = 0
        data = egat_api.get_total_lasted_customerDirect()
        egat_value += data['value']
        total_value += data['value']

        percent = round(egat_value*100/total_value,2)
        item_egat = ItemWithPercent(tag='egat', value=round(egat_value,4), percent=percent)
        percent = round(mea_value*100/total_value,2)
        item_mea = ItemWithPercent(tag='mea', value=round(mea_value,4), percent=percent)
        percent = round(pea_value*100/total_value,2)
        item_pea = ItemWithPercent(tag='pea', value=round(pea_value,4), percent=percent)
        percent = round(exp_value*100/total_value,2)
        item_exp = ItemWithPercent(tag='export', value=round(exp_value,4), percent=percent)
        item_total = ItemWithPercent(tag='total', value=round(total_value,4), percent=100)

        items = [item_total, item_egat, item_mea, item_pea, item_exp]

        logger.info(f'runtime: {runtime() - start_time:3f} s')
        # * Update peak
        try:
            if not is_include_ips:
                peaks = await get_summary_peak(db)
                is_today_exist = False
                for peak in peaks.items:
                    if peak.tag == 'today':
                        is_today_exist = True
                        if peak.value <= item_total.value:
                            peak = schemas.PeakDayCreate(
                                peak_date=data_datetime.date(),
                                peak_datetime=data_datetime,
                                peak_type='demand',
                                value=peak
                            )
                            logger.debug(f'peak: {peak}')
                            await crud.upsert_peak(db, peak=peak)
                if not is_today_exist:
                    await get_profile_demand(db=db)
        except Exception as e:
            await get_profile_demand(db=db)
            logger.error(e)
        return Items(datetime=data_datetime, status='ok', items=items)
    
    except Exception as e:

        logger.info(f'runtime: {runtime() - start_time:3f} s')
        logger.exception(e)
        return Items(datetime=datetime.now(), status='error', items=[])

@router.get("/profile/supply")
async def get_profile_supply(is_include_ips: Optional[bool] = True, source: int = 1, db: AsyncSession = Depends(get_db)) -> Items:
    start_time = runtime()
    items = Items(datetime=datetime.now(), status='ok')
    total_item = TimeseriesItem(tag='total')
    egat_item = TimeseriesItem(tag='egat')
    ipp_item = TimeseriesItem(tag='ipp')
    spp_item = TimeseriesItem(tag='spp')
    vspp_item = TimeseriesItem(tag='vspp')
    ips_item = TimeseriesItem(tag='ips')
    imp_item = TimeseriesItem(tag='imp')

    try:
        #* EGAT, IPP, SPP
        logger.info('retriveing EGAT, IPP, SPP, IMP')
        if source == 1:
            for data in egat_api.get_profile_genMw_group_by_type_source1(items.datetime.date()):
                if  data['tag'] == 'egat':
                    egat_item.values = data['values']
                elif data['tag'] == 'ipp':
                    ipp_item.values = data['values']
                elif data['tag'] == 'spp':
                    spp_item.values = data['values']
                elif data['tag'] == 'imp':
                    imp_item.values = data['values']
                elif data['tag'] == 'total':
                    total_item.values = data['values']
        elif source == 2:
            for data in egat_api.get_profile_genMw_group_by_type_source2(items.datetime.date()):
                if  data['tag'] == 'egat':
                    egat_item.values = data['values']
                elif data['tag'] == 'ipp':
                    ipp_item.values = data['values']
                elif data['tag'] == 'spp':
                    spp_item.values = data['values']
                elif data['tag'] == 'imp':
                    imp_item.values = data['values']
                elif data['tag'] == 'total':
                    total_item.values = data['values']
        else:
            raise HTTPException(status_code=400, detail='soucre should be 1, 2, or 3')
        del data 
        gc.collect()
        # *VSPP
        logger.info('retriveing VSPP')

        start_min = total_item.values[0][0].hour*60 + total_item.values[0][0].minute
        end_min = 24*60
        func2 = iter(egat_api.get_reqMw_profile(14, datetime.today(), strat_min=start_min, end_min=end_min, interval_min=30))
        func3 = iter(egat_api.get_reqMw_profile(15, datetime.today(), strat_min=start_min, end_min=end_min, interval_min=30))
        func4 = iter(egat_api.get_reqMw_profile(16, datetime.today(), strat_min=start_min, end_min=end_min, interval_min=30))
        func5 = iter(egat_api.get_reqMw_profile(12, datetime.today(), strat_min=start_min, end_min=end_min, interval_min=30))
        i = 0
        for data1 in egat_api.get_reqMw_profile(13, datetime.today(), strat_min=start_min, end_min=end_min, interval_min=30):
            data2 =next(func2)
            data3 =next(func3)
            data4 =next(func4)
            data5 =next(func5)
            value = data1[1]+data2[1]+data3[1]+data4[1]+data5[1]
            vspp_item.values.append((total_item.values[i][0], value))
            old_dt, old_value = total_item.values[i]
            total_item.values[i]= (old_dt, old_value+value)
            i += 1
        

        # *IPS
        logger.info('retriveing IPS')
        if is_include_ips:
            datetime_from = total_item.values[0][0]
            datetime_to = total_item.values[-1][0]
            i = 0
            async for data in crud.get_profile_dummy_data(db=db, category='ips', datetime_from=datetime_from, datetime_to=datetime_to, min_interval=30):
                ips_item.values.append(data)
                old_dt, old_value = total_item.values[i]
                total_item.values[i]= (old_dt, old_value+data[1])
                i += 1

        items.items = [total_item, egat_item, ipp_item, spp_item, vspp_item, ips_item, imp_item]
    except Exception as e:
        logger.exception(e)
        items.status = 'error'
    logger.info(f'runtime: {runtime() - start_time:3f} s')
    return items


@router.get("/profile/supply/fuel")
async def get_profile_supply_fuel(is_include_ips: Optional[bool] = True, source: int = 1, db: AsyncSession = Depends(get_db)) -> Items:
    start_time = runtime()
    items = Items(datetime=datetime.now(), status='ok')
    gas_item = TimeseriesItem(tag='ก๊าซธรรมชาติ')
    renew_item = TimeseriesItem(tag='พลังงานทดแทน')
    hydro_item = TimeseriesItem(tag='พลังงานน้ำ')
    coal_item = TimeseriesItem(tag='ถ่านหิน')
    oil_item = TimeseriesItem(tag='น้ำมัน')

    try:
        #* EGAT, IPP, SPP
        logger.info('retriveing EGAT, IPP, SPP, IMP')
        if source == 1:
            for data in egat_api.get_profile_genMw_group_by_fuel_source1(items.datetime.date()):
                if  data['tag'] == 'ก๊าซธรรมชาติ':
                    gas_item.values = data['values']
                elif data['tag'] == 'พลังงานทดแทน':
                    renew_item.values = data['values']
                elif data['tag'] == 'พลังงานน้ำ':
                    hydro_item.values = data['values']
                elif data['tag'] == 'ถ่านหิน':
                    coal_item.values = data['values']
                elif data['tag'] == 'น้ำมัน':
                    oil_item.values = data['values']
        elif source == 2:
            for data in egat_api.get_profile_genMw_group_by_fuel_source2(items.datetime.date()):
                if  data['tag'] == 'ก๊าซธรรมชาติ':
                    gas_item.values = data['values']
                elif data['tag'] == 'พลังงานทดแทน':
                    renew_item.values = data['values']
                elif data['tag'] == 'พลังงานน้ำ':
                    hydro_item.values = data['values']
                elif data['tag'] == 'ถ่านหิน':
                    coal_item.values = data['values']
                elif data['tag'] == 'น้ำมัน':
                    oil_item.values = data['values']
        else:
            raise HTTPException(status_code=400, detail='soucre should be 1, 2, or 3')
        del data 
        gc.collect()
        
        datetime_from = gas_item.values[0][0]
        datetime_to = gas_item.values[-1][0]
        i = 0
        renew_list = ['PEA ผลิต', 'ขยะ', 'ชีวภาพ', 'ชีวมวล', 'พพ. ผลิต', 'พลังงานความร้อนเหลือ', 'ลม', 'แสงอาทิตย์']
        gas_list = ['ก๊าซธรรมชาติ', 'co-gen']
        
        # *VSPP
        logger.info('retriveing VSPP')
        async for fuel, data in crud.get_profile_dummy_data_grouped_by_value_tag(db=db, category='vspp', datetime_from=datetime_from, datetime_to=datetime_to, min_interval=30):
            if fuel in gas_list:
                old_dt, old_value = gas_item.values[i]
                gas_item.values[i]= (old_dt, old_value+data[1])
                i += 1
            elif fuel in renew_list:
                old_dt, old_value = renew_item.values[i]
                renew_item.values[i]= (old_dt, old_value+data[1])
                i += 1
            elif fuel == 'ถ่านหิน':
                old_dt, old_value = coal_item.values[i]
                coal_item.values[i]= (old_dt, old_value+data[1])
                i += 1
            elif fuel == 'พลังงานน้ำ':
                old_dt, old_value = hydro_item.values[i]
                hydro_item.values[i]= (old_dt, old_value+data[1])
                i += 1

        # *IPS
        logger.info('retriveing IPS')
        if is_include_ips:
            async for fuel, data in crud.get_profile_dummy_data_grouped_by_value_tag(db=db, category='vspp', datetime_from=datetime_from, datetime_to=datetime_to, min_interval=30):
                if fuel in gas_list:
                    old_dt, old_value = gas_item.values[i]
                    gas_item.values[i]= (old_dt, old_value+data[1])
                    i += 1
                elif fuel in renew_list:
                    old_dt, old_value = renew_item.values[i]
                    renew_item.values[i]= (old_dt, old_value+data[1])
                    i += 1
                elif fuel == 'ถ่านหิน':
                    old_dt, old_value = coal_item.values[i]
                    coal_item.values[i]= (old_dt, old_value+data[1])
                    i += 1
                elif fuel == 'พลังงานน้ำ':
                    old_dt, old_value = hydro_item.values[i]
                    hydro_item.values[i]= (old_dt, old_value+data[1])
                    i += 1

        items.items = [gas_item, coal_item, oil_item, hydro_item, renew_item]
    except Exception as e:
        logger.exception(e)
        items.status = 'error'
    logger.info(f'runtime: {runtime() - start_time:3f} s')
    return items


@router.get("/profile/demand") #! get profile and update peak that day
async def get_profile_demand( profile_date: date = date.today(), is_update_peak: bool = True,
db: AsyncSession = Depends(get_db)) -> Items:
    start_time = runtime()
    items = Items(datetime=datetime.now(), status='ok')
    try:
    # *get profile
    # *vspp, Export, 3E
        data1 = egat_api.get_reqMw(1, profile_date)['list']
        data2 = egat_api.get_reqMw(2, profile_date)['list']
        data3 = egat_api.get_reqMw(3, profile_date)['list']
        data4 = egat_api.get_reqMw(4, profile_date)['list']
        data5 = egat_api.get_reqMw(5, profile_date)['list']
        data6 = egat_api.get_reqMw(7, profile_date)['list']
        data7 = egat_api.get_reqMw(7, profile_date)['list']
        data12 = egat_api.get_reqMw(12, profile_date)['list']
        data13 = egat_api.get_reqMw(13, profile_date)['list']
        data14 = egat_api.get_reqMw(14, profile_date)['list']
        data15 = egat_api.get_reqMw(15, profile_date)['list']
        data16 = egat_api.get_reqMw(16, profile_date)['list']

        data_timestamp = datetime.combine(profile_date, time(0, 0))
        get_customerDirect = iter(egat_api.get_total_customerDirect(start_date=profile_date, end_date=profile_date))
        value_customerDirect = 0
        item = TimeseriesItem(tag='actual')
        peak_value = 0
        peak_data_timestamp = data_timestamp

        for i in range(len(data1)):
            value = 0
            hour = floor(data1[i][0]/3600)
            min = floor((data1[i][0] - hour*3600)/60)
            if hour < 24:
                data_timestamp = data_timestamp.replace(hour=hour,minute=min)
            else:
                data_timestamp = data_timestamp.replace(hour=0,minute=0)
                data_timestamp += timedelta(days=1)

            
            # *get egat direccustomer
            try:
                if min%30 == 0:
                    data = next(get_customerDirect)
                    if data and 'value' in data:
                        value_customerDirect = data['value']  # Update only if new data is available
            except StopIteration:
                pass  # Keep previous value_customerDirect as is
            
            value += data1[i][-1]
            value += data2[i][-1]
            value += data3[i][-1]
            value += data4[i][-1]
            value += data5[i][-1]
            value += data6[i][-1]
            value += data7[i][-1]
            value += data12[i][-1]
            value += data13[i][-1]
            value += data14[i][-1]
            value += data15[i][-1]
            value += data16[i][-1]
            value += value_customerDirect
            value = round(value, 4)
            # *define peak
            if value >= peak_value: 
                peak_value = value
                peak_data_timestamp = data_timestamp

            item.values.append((data_timestamp, value))


        items.items = [item]
    
    # *updaet peak acording peak date
        if is_update_peak:
            peak = schemas.PeakDayCreate(
                peak_date=peak_data_timestamp.date(),
                peak_datetime=peak_data_timestamp,
                peak_type='demand',
                value=peak_value
            )
            logger.debug(f'peak: {peak}')
            await crud.upsert_peak(db, peak=peak)
    except Exception as e:
        items.status = 'error'
        logger.exception(e)

    logger.info(f'runtime: {runtime() - start_time:3f} s')
    return items

@router.get("/summary/peak")
async def get_summary_peak(db: AsyncSession = Depends(get_db)):
    start_time = runtime()

    peaks = await crud.get_summary_peak(db=db, peak_type="demand")

    items = []
    for tag, peak in peaks.items():
        if peak:
            logger.debug(peak)
            items.append(ItemWithTimestamp(
                tag=tag,
                value=peak["value"],
                timestamp=peak["timestamp"]
            ))
    logger.info(f'runtime: {runtime() - start_time:3f} s')

    return Items(
        datetime=datetime.now(),
        status="ok",
        items=items
    )

@router.get("/cont/project/renew")
async def get_count_project_renew(db: AsyncSession = Depends(get_db)) -> Items:
    start_time = runtime()
    items = Items(datetime=datetime.now(), status='ok')

    fuel_map_th = {
        'Solar': 'พลังงานแสงอาทิตย์',
        'Hydro': 'หลังงานน้ำ',
        'Wind': 'หลังงานลม',
        'Biogas': 'ก๊าซชีวภาพ',
        'Biomass': 'ชีวมวล',
        'Waste': 'ขยะ',
        'Geothermal': 'พลังงานความร้อนใต้พิภพ',
        'RE - Others': 'อื่นๆ',
        'N/A': 'ไม่ระบุบ',
    }
    total_count = 0
    for fuel_en in fuel_map_th.keys():
        count = await crud.count_active_projects_by_fuel(db=db, fuel=fuel_en, fuel_type='renew')
        total_count += count
        items.items.append(Item(
            tag=fuel_map_th[fuel_en], value=count
        ))
    items.items.append(Item(tag='total',value=total_count))
    logger.info(f'runtime: {runtime() - start_time:3f} s')
    return items

@router.get("/cont/project/fossil")
async def get_count_project_fossil(db: AsyncSession = Depends(get_db)) -> Items:
    start_time = runtime()
    items = Items(datetime=datetime.now(), status='ok')

    fuel_map_th = {
        'Natural Gas': 'ก๊าซธรรมชาติ',
        'Diesel Oil': 'น้ำมันดีเซล',
        'Bunker Oil': 'น้ำมันบังเกอร์',
        'Pitch': 'พิช',
        'Lignite': 'ถ่านหินลิกไนต์',
        'Coal': 'ถ่านหิน',
    }
    total_count = 0
    for fuel_en in fuel_map_th.keys():
        count = await crud.count_active_projects_by_fuel(db=db, fuel=fuel_en, fuel_type='fossil')
        total_count += count
        items.items.append(Item(
            tag=fuel_map_th[fuel_en], value=count
        ))
    items.items.append(Item(tag='total',value=total_count))
    logger.info(f'runtime: {runtime() - start_time:3f} s')
    return items

@router.get("/project/location/fuel")
async def get__project_location_by_fuel(db: AsyncSession = Depends(get_db)) -> Items:
    start_time = runtime()
    items = Items(datetime=datetime.now(), status='ok')
    try:
        async for row in crud.get_projects_location_by_fuel(db=db):
            items.items.append(LocationItem(tag=row['fuel'], lat=row['lat'], lng=row['lng']))
    
    except Exception as e:
        logger.exception(e)
    logger.info(f'runtime: {runtime() - start_time:3f} s')
    return items

@router.post("/submit/dummy")
async def submit_dummy_data(title: str,
                            file: UploadFile = File(...),
                            db: AsyncSession = Depends(get_db)):
    start_time = runtime()
    submit_timestamp = datetime.now()

    # *category validation
    valid_titles = ['ips','vspp']
    if title not in valid_titles:
        raise HTTPException(
            status_code=400,
            detail=f'tile is not valid. valid list {valid_titles}')

    # *file validation
    try:
        content = await file.read()
        decoded = content.decode("utf-8-sig")
        csv_reader = csv.DictReader(StringIO(decoded))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'invalidate file')

    REQUIRED_COLUMNS = {'DTM', 'ZONE', 'PROVINCE', 'TYPE', 'VALDUMMY'}
    actual_columns = set(csv_reader.fieldnames or [])
    logger.debug(actual_columns)
    missing = REQUIRED_COLUMNS - actual_columns
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}")
    # Data Process rows (example: collect to return)
    error_logs = []
    processed_count = 0
    csv_reader = list(csv.DictReader(StringIO(decoded)))
    total_count = len(csv_reader)

    for row in csv_reader:
        try:
            data = schemas.DummyDataCreate(
                submit_timestamp=submit_timestamp,
                data_timestamp=datetime.fromisoformat(row['DTM']),
                category=title,
                zone=row['ZONE'],
                province=row['PROVINCE'],
                value_tag=row['TYPE'],
                value=float(row['VALDUMMY']))
            await crud.create_dummy_data(db, data)

        except Exception as e:
            msg = f'{e}, data:{row}'
            logger.error(msg)
            error_logs.append(msg)
        processed_count += 1
        logger.info(
            f"Processed {processed_count} / {total_count} records for title='{title}'"
        )

    logger.info(f'runtime: {runtime() - start_time:3f} s')

    if len(error_logs) == 0:
        return Msg(status='ok', message='insert without error')
    else:
        return Msg(status='error',
                           message=f'errors: {"\n".join(error_logs) }')

@router.post("/submit/peak")
async def submit_peak(title: str,
                            file: UploadFile = File(...),
                            db: AsyncSession = Depends(get_db)):
    start_time = runtime()

    # *category validation
    valid_titles = ['demand']
    if title not in valid_titles:
        raise HTTPException(
            status_code=400,
            detail=f'tile is not valid. valid list {valid_titles}')

    # *file validation
    try:
        content = await file.read()
        decoded = content.decode("utf-8-sig")
        csv_reader = csv.DictReader(StringIO(decoded))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'invalidate file')

    REQUIRED_COLUMNS = {'Time', 'Value'}
    actual_columns = set(csv_reader.fieldnames or [])
    logger.debug(actual_columns)
    missing = REQUIRED_COLUMNS - actual_columns
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}")
    # Data Process rows (example: collect to return)
    error_logs = []
    processed_count = 0
    csv_reader = list(csv.DictReader(StringIO(decoded)))
    total_count = len(csv_reader)

    for row in csv_reader:
        try:
            data_timestamp = datetime.fromisoformat(row['Time'])
            data = schemas.PeakDayCreate(
                peak_date=data_timestamp.date(),
                peak_datetime=data_timestamp,
                peak_type=title,
                value=float(row['Value']))
            await crud.upsert_peak(db, data)

        except Exception as e:
            msg = f'{e}, data:{row}'
            logger.error(msg)
            error_logs.append(msg)
        processed_count += 1
        logger.info(
            f"Processed {processed_count} / {total_count} records for title='{title}'"
        )

    logger.info(f'runtime: {runtime() - start_time:3f} s')

    if len(error_logs) == 0:
        return Msg(status='ok', message='insert without error')
    else:
        return Msg(status='error',
                           message=f'errors: {"\n".join(error_logs) }')

@router.post("/submit/project")
async def submit_project_data(file: UploadFile = File(...),
                              db: AsyncSession = Depends(get_db)):
    start_time = runtime()
    submit_timestamp = datetime.now()

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400,
                            detail="Only Excel files are supported")

    try:
        content = await file.read()
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
        sheet = workbook.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")

    headers = [
        cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))
    ]
    header_index = {h: i for i, h in enumerate(headers)}

    processed_count = 0
    error_logs = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data = {h: row[header_index[h]] for h in headers}
        try:
            project = schemas.ProjectCreate(
                g_project_key=str(data.get("G_PROJECT_KEY") or ""),
                spp_vspp_rowid=str(data.get("SPP_VSPP_ROWID") or ""),
                e_license_rowid=str(data.get("E_LICENSE_ROWID") or ""),
                pk_powersystemresource=str(
                    data.get("PK_POWERSYSTEMRESOURCE") or ""),
                erc_cd=str(data.get("ERC_CD") or ""),
                org=data.get("ORG"),
                spp_vspp_plant_cd=data.get("SPP_VSPP_PLANT_CD"),
                ppa_contract_no=data.get("PPA_CONTRACT_NO"),
                project_name=data.get("PROJECT_NAME"),
                spp_vspp_project_name=data.get("SPP_VSPP_PROJECT_NAME"),
                licensee=data.get("LICENSEE"),
                display_addr=data.get("DISPLAY_ADDR"),
                subdistrict=data.get("SUBDISTRICT"),
                district=data.get("DISTRICT"),
                province=data.get("PROVINCE"),
                country_zone=data.get("COUNTRY_ZONE"),
                egat_zone=data.get("EGAT_ZONE"),
                contract_status=data.get("CONTRACT_STATUS"),
                e_license_instl_mw=data.get("E_LICENSE_INSTL_MW"),
                e_license_instl_kva=data.get("E_LICENSE_INSTL_KVA"),
                installed_cap_mw=data.get("INSTALLED_CAP_MW"),
                contracted_cap_mw=data.get("CONTRACTED_CAP_MW"),
                project_type=data.get("PROJECT_TYPE"),
                contract_type=data.get("CONTRACT_TYPE"),
                technology_a_group_1=data.get("TECHNOLOGY_A_GROUP_1"),
                technology_a_group_2=data.get("TECHNOLOGY_A_GROUP_2"),
                technology_a_detail=data.get("TECHNOLOGY_A_DETAIL"),
                primary_fuel_a_group_1=data.get("PRIMARY_FUEL_A_GROUP_1"),
                primary_fuel_a_group_2=data.get("PRIMARY_FUEL_A_GROUP_2"),
                primary_fuel_a_group_3=data.get("PRIMARY_FUEL_A_GROUP_3"),
                secondary_fuel_a_group_1=data.get("SECONDARY_FUEL_A_GROUP_1"),
                secondary_fuel_a_group_2=data.get("SECONDARY_FUEL_A_GROUP_2"),
                secondary_fuel_a_group_3=data.get("SECONDARY_FUEL_A_GROUP_3"),
                technology_b_group_1=data.get("TECHNOLOGY_B_GROUP_1"),
                technology_b_group_2=data.get("TECHNOLOGY_B_GROUP_2"),
                technology_b_detail=data.get("TECHNOLOGY_B_DETAIL"),
                primary_fuel_b_group_1=data.get("PRIMARY_FUEL_B_GROUP_1"),
                primary_fuel_b_group_2=data.get("PRIMARY_FUEL_B_GROUP_2"),
                primary_fuel_b_group_3=data.get("PRIMARY_FUEL_B_GROUP_3"),
                secondary_fuel_b_group_1=data.get("SECONDARY_FUEL_B_GROUP_1"),
                secondary_fuel_b_group_2=data.get("SECONDARY_FUEL_B_GROUP_2"),
                secondary_fuel_b_group_3=data.get("SECONDARY_FUEL_B_GROUP_3"),
                utilization=data.get("UTILIZATION"),
                scod=data.get("SCOD"),
                cod=data.get("COD"),
                lat=data.get("LAT"),
                lng=data.get("LNG"),
                is_egat_sys_gen=(str(data.get("IS_EGAT_SYS_GEN")) == '1'),
                is_sharing_lic=(str(data.get("IS_SHARING_LIC")) == '1'),
                licensingno=data.get("LICENSINGNO"),
                erc_district_no=str(data.get("ERC_DISTRICT_NO") or ""),
                erc_district_displayname=data.get("ERC_DISTRICT_DISPLAYNAME"),
                month_key=str(data.get("MONTH_KEY") or ""),
                update_timestamp=submit_timestamp)
            await crud.upsert_project(db, project)
        except Exception as e:
            if db.in_transaction():
                await db.rollback()
            error_logs.append(f"{e}, data: {data}")
            logger.error(f"{e}, data: {data}")
        processed_count += 1
        logger.info(f"Processed {processed_count} / {sheet.max_row - 1}")

    logger.info(f'runtime: {runtime() - start_time:3f} s')
    
    if len(error_logs) == 0:
        return Msg(status='ok', message='insert without error')
    else:
        return Msg(
            status='error',
            message=f'errors: {len(error_logs)} records failed.')


